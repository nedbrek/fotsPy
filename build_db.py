#!/usr/bin/env python3
import fots_db
import fots_lib as fots
import glob
import openpyxl
import sqlite3

### main
if __name__ == '__main__':
    import argparse
    arg_parser = argparse.ArgumentParser()
    arg_parser.add_argument("turnfiles", nargs=1, help = "Path to Excel turn files")
    arg_parser.add_argument("prefix", nargs=1, help = "Common prefix on Excel turn files")
    arg_parser.add_argument("dbfile", nargs=1, help = "DB file to create")
    args = arg_parser.parse_args()

    path = "{}/{}".format(args.turnfiles[0], args.prefix[0])
    files = glob.glob('{}*.xls*'.format(path))
    sfiles = sorted(files, key=lambda x: int("".join([i for i in x if i.isdigit()])))

    conn = sqlite3.connect(args.dbfile[0])
    cur = conn.cursor()

    print("Processing {}".format(sfiles[0]))
    wb_obj = openpyxl.load_workbook(sfiles[0])

    data = fots.Fots()

    map_sheet = wb_obj["Map"]
    data.buildStarmap(map_sheet)

    col_sheet = wb_obj["Colonies"]
    data.buildColonies(col_sheet)

    fots_db.buildDb(conn)
    fots_db.insertStarmap(conn, data)

    habs = fots.Habs()
    habs.readExcel(wb_obj)
    fots_db.insertHabs(conn, habs)

    summary_sheet = wb_obj["Summary"]
    row = 7
    col = 12
    turn = summary_sheet.cell(row, col).value

    fots_db.insertSurveys(conn, data, turn)

    for excel_file in sfiles[1:]:
        print("Processing {}".format(excel_file))
        wb_obj = openpyxl.load_workbook(excel_file)
        data = fots.Fots()

        map_sheet = wb_obj["Map"]
        data.buildStarmap(map_sheet)

        col_sheet = wb_obj["Colonies"]
        data.buildColonies(col_sheet)

        summary_sheet = wb_obj["Summary"]
        row = 7
        col = 12
        turn = summary_sheet.cell(row, col).value

        fots_db.insertSurveys(conn, data, turn)

