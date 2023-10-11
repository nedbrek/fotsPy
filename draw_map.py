#!/usr/bin/env python3

from enum import Enum, auto
import fots_lib as fots
import fots_db
import json
import openpyxl
import re
import sqlite3
import sys
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog

def copySysTree(event):
#{
    w = event.widget
    root.clipboard_clear()
    items = w.selection()
    for i in items:
        parent = w.parent(i)
        if parent == '':
            root.clipboard_append(w.item(i, 'text'))
            continue

        system = w.item(parent, 'text')
        pnum = w.item(i, 'text')
        vals = w.item(i, 'values')
        root.clipboard_append("{} #{:02d}\t{}\n".format(system, pnum, '\t'.join(vals)))
#}

### types
## State machine for system classification
class EHabs(Enum):
    NO_HABS   = auto()  # nothing worth colonizing
    ALL_HABS  = auto()  # all habitable worlds colonized
    GOOD_HABS = auto()  # good colony site
    BAD_HABS  = auto()  # marginal colony site

### globals
## grid size (pixels)
side = 10 # length of the side of a square
zoom_levels = [
   10,
   20,
   40,
   60,
   90,
   150
]

data = None
hab_data = dict()
canvas = None
text = None

### functions
def selectStar(event) -> None:
    """ Respond to click on a system """
    #{
    global canvas
    tags = canvas.gettags('current')
    regex = re.compile("key_(.*)")
    tag = list(filter(regex.match, tags))
    key = regex.match(tag[0]).group(1)
    base_tag = "key_{}".format(key)

    # delete old info
    global system_tree
    for s in system_tree.get_children(''):
        system_tree.delete(s)

    # restore border on old
    canvas.itemconfigure('active', outline='black', width=1)
    canvas.dtag('active')

    canvas.addtag('active', 'withtag', base_tag)
    canvas.itemconfigure('active', outline='red', width=3)

    global conn
    cur = conn.cursor()

    sys = system_tree.insert('', 'end', text=key)
    system_tree.item(sys, open=True)

    res = cur.execute("""SELECT
        num, type, rp, srp, owner, turn
        FROM surveys a
        WHERE key=? AND turn=(
            SELECT max(turn) FROM surveys
            WHERE key=? AND num=a.num
        )
        ORDER BY num
    """, (key, key)).fetchall()
    for v in res:
        num = v[0]
        wtype = v[1]
        rp = v[2]
        srp = v[3]
        owner = v[4]
        turn = v[5]

        if srp == 0:
            srp = ''

        hab = cur.execute("SELECT val FROM hab_data WHERE key=?", (wtype,)).fetchone()[0]

        system_tree.insert(sys, 'end', text=num, values=(hab, rp, srp, owner, turn))
    #}


def drawDb() -> None:
    """ Read the database and draw it """
    #{
    global canvas
    global conn
    global data
    global hab_data
    global side

    canvas.delete('all')

    cur = conn.cursor()
    first_x = int(getDbVal(cur, "first_x"))
    first_y = int(getDbVal(cur, "first_y"))

    # pull data from db
    stars = cur.execute('SELECT gridx, gridy, type FROM starmap').fetchall()
    for row in stars:
        xoff = row[0]
        yoff = row[1]
        x = (xoff - first_x) * side
        y = (yoff - first_y) * side
        star_type = row[2]

        fill_color = 'black'
        grid_res = cur.execute("""SELECT color
            FROM comm_grid WHERE gridx=? AND gridy=?
            ORDER BY turn DESC LIMIT 1
        """, (xoff, yoff)).fetchone()
        if grid_res is not None:
            fill_color = "#{}".format(grid_res[0][2:])

        key = "key_{}".format(fots.Fots.makeKey(xoff, yoff))
        tags = [key, 'star']
        item_id = canvas.create_rectangle(x, y, x + side, y + side, fill=fill_color, tags=tags)

        # TODO db doesn't have star last survey...
        if side >= 40:
            fill_color = 'grey'

            # hasattr(s, 'last_survey'):
            if False:
                state = EHabs.NO_HABS
                for w in s.worlds:
                    hab = hab_data[w.type]
                    own = hasattr(w, 'owner')
                    if own:
                        # TODO show room for expansion
                        state = EHabs.ALL_HABS
                        fill_color = 'white'
                    elif state == EHabs.NO_HABS and (hab == 'M' or hab == 'N'):
                        state = EHabs.GOOD_HABS
                        fill_color = '#00C000'
                    elif state == EHabs.NO_HABS and hab == 'P':
                        state = EHabs.BAD_HABS
                        fill_color = 'red'
                #foreach world
            #if survey data

            canvas.create_text(x + side / 2, y + side / 2, text=star_type, fill=fill_color, tags=['icon'])
        #if text is visible
    #foreach star

    canvas.configure(scrollregion = canvas.bbox('all'))
    canvas.tag_bind('star', "<1>", selectStar)
#}def drawDb

def zoomOut(event) -> None:
    """ Reduce zoom level and redraw """
    global side
    global zoom_levels

    idx = zoom_levels.index(side)
    if idx == 0:
        return
    side = zoom_levels[idx - 1]
    drawDb()

def zoomIn(event) -> None:
    """ Increase zoom level and redraw """
    global side
    global zoom_levels

    idx = zoom_levels.index(side)
    if idx + 1 == len(zoom_levels):
        return
    side = zoom_levels[idx + 1]
    drawDb()

### db functions
def addExcel(root) -> None:
    """ Ask user for an Excel file to read into the db """
    #{
    global conn

    filename = filedialog.askopenfilename()
    if not filename:
        return

    wb_obj = openpyxl.load_workbook(filename)

    map_sheet = wb_obj["Map"]
    data = fots.Fots()
    data.buildStarmap(map_sheet)

    col_sheet = wb_obj["Colonies"]
    data.buildColonies(col_sheet)

    summary_sheet = wb_obj["Summary"]
    row = 7
    col = 12
    turn = summary_sheet.cell(row, col).value

    fots_db.insertSurveys(conn, data, turn)

    drawDb()
    # TODO pull database name
    root.winfo_toplevel().title("Fots turn {}".format(turn))
    #}

def getDbVal(cur, key, col="val", table="notes"):
    res = cur.execute("SELECT {} FROM {} WHERE key=?".format(col, table), (key,))
    t = res.fetchone()
    if t is None:
        return None
    return t[0]

def findSystem(stars, key):
    for s in stars:
        if fots.Fots.makeKey(s.xoff, s.yoff) == key:
            return s
    return None

def tempCode():
    for row in cur.execute('SELECT key, type, rp, srp, owner, turn FROM surveys'):
        s = findSystem(data.stars, row[0])
        w = fots.World()
        w.type  = row[1]
        w.rp    = row[2]
        w.srp   = row[3]
        w.owner = row[4]
        w.last_survey = row[5]

        s.worlds.append(w)

def fillFots(cur):
#{
    global data
    global hab_data

    data = fots.Fots()

    # pull habs
    hab_data = fots.Habs()
    for row in cur.execute('SELECT key, val FROM hab_data'):
        hab_data.habs[row[0]] = row[1]
#}

def openDb(root):
    """ Open an existing database """
    #{
    filename = filedialog.askopenfilename()
    if not filename:
        return

    global conn
    conn = sqlite3.connect(filename)
    cur = conn.cursor()
    fillFots(cur)

    drawDb()
    res = cur.execute("SELECT MAX(turn) FROM comm_grid").fetchone()
    turn = "<no surveys>" if res is None else res[0]
    root.winfo_toplevel().title("Fots {} turn {}".format(filename, turn))
    #}

def newDb(root):
    """ Create a database and load an Excel turn """
    #{
    filename = filedialog.asksaveasfilename()
    if not filename:
        return

    excel_file = filedialog.askopenfilename()
    if not excel_file:
        return

    global conn
    conn = sqlite3.connect(filename)

    wb_obj = openpyxl.load_workbook(excel_file)

    map_sheet = wb_obj["Map"]
    global data
    data = fots.Fots()
    data.buildStarmap(map_sheet)

    # create tables
    fots_db.buildDb(conn)

    # populate starmap
    fots_db.insertStarmap(conn, data)

    habs = fots.Habs()
    habs.readExcel(wb_obj)
    fots_db.insertHabs(conn, habs)

    summary_sheet = wb_obj["Summary"]
    row = 7
    col = 12
    turn = summary_sheet.cell(row, col).value

    fots_db.insertSurveys(conn, data, turn)

    drawDb()
    root.winfo_toplevel().title("Fots {} turn {}".format(filename, turn))
    #}

### main
if __name__ == '__main__':
    import argparse
    arg_parser = argparse.ArgumentParser()
    # TODO add param for db and excel
    args = arg_parser.parse_args()

    # build the GUI
    root = tk.Tk()
    root.option_add('*tearOff', False)

    root.winfo_toplevel().title("Fots - <no db loaded>")

    ## menu bar
    menubar = tk.Menu(root)
    root['menu'] = menubar

    ### file menu
    menu_file = tk.Menu(menubar)
    menubar.add_cascade(menu=menu_file, label='File')
    menu_file.add_command(label='New', command=lambda: newDb(root))
    menu_file.add_command(label='Open', command=lambda: openDb(root))
    menu_file.add_command(label='Add Report', command=lambda: addExcel(root))
    menu_file.add_command(label='Exit', command=sys.exit)

    ## gui elements
    pane = ttk.PanedWindow(root, orient='horizontal')
    pane.pack(expand=True, fill='both')

    right_frame = ttk.Frame()

    left_pane = ttk.PanedWindow(root, orient='vertical', width=385)

    system_tree = ttk.Treeview()
    system_tree['columns'] = ('hab', 'rp', 'srp', 'owner', 'turn')
    system_tree.column('#0', width=95)
    system_tree.column('hab', width=15)
    system_tree.column('rp', width=65)
    system_tree.column('srp', width=55)
    system_tree.column('owner', width=60)
    system_tree.column('turn', width=35)
    left_pane.add(system_tree)

    text = tk.Text()
    left_pane.add(text)

    pane.add(left_pane)
    pane.add(right_frame)

    # star map
    hsb = ttk.Scrollbar(right_frame, orient=tk.HORIZONTAL)
    vsb = ttk.Scrollbar(right_frame, orient=tk.VERTICAL)

    canvas = tk.Canvas(right_frame, yscrollcommand = vsb.set, xscrollcommand = hsb.set)

    hsb['command'] = canvas.xview
    vsb['command'] = canvas.yview

    hsb.pack(fill='x', side='bottom')
    vsb.pack(fill='y', side='right')
    canvas.pack(expand=True, fill='both', side='right')

    system_tree.bind('<Control-c>', copySysTree)

    # bind zoom
    root.bind('<minus>',       zoomOut)
    root.bind('<KP_Subtract>', zoomOut)
    root.bind('<plus>',        zoomIn)
    root.bind('<KP_Add>',      zoomIn)

    root.mainloop()

