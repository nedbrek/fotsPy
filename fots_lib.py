import openpyxl
import re
import sys

### classes
class World:
    def __init__(self):
        self.type = "<error>"
        self.rp = -1
        self.srp = -1

def compareWorlds(key, w1, w2):
    if w1.type != w2.type:
        print("Type mismatch {} {} {}".format(key, w1.type, w2.type))

    if w1.rp != w2.rp:
        print("RP mismatch {} {} {}".format(key, w1.rp, w2.rp))

    if w1.srp != w2.srp:
        print("SRP mismatch {} {} {}".format(key, w1.srp, w2.srp))

    return # TODO check owner once we resolve the formula
    if hasattr(w1, 'owner'):
        if hasattr(w2, 'owner'):
            if w1.owner != w2.owner:
                print("Owner mismatch {} {} {}".format(key, w1.owner, w2.owner))
        else:
            print("Owner2 missing {} {}".format(key, w1.owner))
    elif hasattr(w2, 'owner'):
        print("Owner1 missing {} {}".format(key, w2.owner))

class System:
    def __init__(self):
        self.xoff = -1
        self.yoff = -1
        self.key = "<error>"
        self.star_type = "<error>"
        self.worlds = list()
        self.grid = None

class Habs:
    def __init__(self):
        self.habs = dict()

    def readExcel(self, wb_obj):
        colony_sheet = wb_obj["Colonies"]
        row = 2
        while 1:
            col = 64
            key_val = colony_sheet.cell(row, col).value
            if key_val == None:
                break

            col = 66
            hab_val = colony_sheet.cell(row, col).value
            if hab_val == None:
                print("Error in hab val, row:", row)
                hab_val = "?"

            self.habs[key_val] = hab_val
            row = row + 1

class Fots:
    def __init__(self):
        self.stars = list()

    # excel util
    def resolveFormula(sheet, cell):
        val = cell.value
        if type(val) is int:
            return val

        if val[0] == '=':
            plus_idx = val.find("+")
            idx = val[1:plus_idx]
            v2 = sheet[idx].value
            const = int(val[plus_idx+1:])
            v2 = v2 + const
            cell.value = v2

        return v2

    # find x dimensions
    def findX(map_sheet):
        row = 2
        col = 7
        first_x = map_sheet.cell(row, col).value
        while 1:
            cell = map_sheet.cell(row, col)
            val = cell.value
            if val == None:
                break

            last_x = Fots.resolveFormula(map_sheet, cell)

            col = col + 10
        #end while
        return (first_x, last_x)

    # find grid y
    def findY(map_sheet):
        row = 8
        col = 1
        first_y = map_sheet.cell(row, col).value
        while 1:
            cell = map_sheet.cell(row, col)
            val = cell.value
            if val == None:
                break

            last_y = Fots.resolveFormula(map_sheet, cell)

            row = row + 10
        #end while
        return (first_y, last_y)

    def makeKey(grid_x, grid_y):
        dec_x = int(grid_x / 10)
        sec_x = (grid_x % 10) + 1

        dec_y = int(grid_y / 10)
        sec_y = grid_y % 10
        key = "{}x{}/{:02d}".format(dec_x, dec_y, sec_x + sec_y * 10)
        return key

    def parseComments(self, s, comments):
        last_comment = comments[-1]
        lc_last_col = last_comment.split()
        last_survey = -1
        if (len(lc_last_col) > 2):
            last_survey = int(lc_last_col[2])

        for c in comments:
        #{
            regex = re.compile("Special: Wormhole")
            if regex.search(c):
                # TODO log wormhole
                print("NedW {} '{}'".format(s.key, c))
                continue

            regex = re.compile(": Foreign Fleet")
            m = regex.search(c)
            if m:
                # TODO log foreign fleet
                c = c[0:m.span()[0]]

            cols = c.split()

            if cols[0] == "Last" and cols[1] == "Survey:":
                # TODO? check survey stamps?
                continue

            if cols[1][0] != '#':
                #print("NedS {} '{}'".format(s.key, c))
                continue

            num_cols = len(cols)
            if num_cols == 2:
                print("NedC2 {} '{}'".format(s.key, c))
                continue

            w = World()

            if num_cols == 6:
                w.pnum = int(cols[1][1:])
                w.type = cols[2]
                w.rp = int(cols[3])
                w.srp = 0
                last_survey = int(cols[4])
                if cols[5] != "!":
                    print("NedC6 {} '{}'".format(s.key, c))
            else:
            #{
                #print("Ned {} '{}'".format(s.key, c))
                w.pnum = int(cols[1][1:])
                w.type = cols[2]

                # parse planet tags
                i = 3
                if i >= len(cols):
                    print("NedD {} '{}'".format(s.key, c))
                    continue

                if cols[i] == "Waterworld":
                    # TODO log waterworld
                    i = i + 1

                if cols[i] == "gas":
                    # TODO log gas giant
                    i = i + 1
                    if cols[i] != "giant":
                        print("NedG {} '{}'".format(s.key, c))
                        continue
                    i = i + 1
                    if cols[i] == "-":
                        i = i + 1

                if cols[i] == "asteroids":
                    # TODO log asteroids
                    i = i + 1

                if cols[i] == "oort":
                    # TODO log oort cloud
                    i = i + 1
                    if cols[i] != "cloud":
                        print("NedO {} '{}'".format(s.key, c))
                        continue
                    i = i + 1

                if cols[i] == "dextro" or cols[i] == "levo" or cols[i] == "silicon" or cols[i] == "methane":
                    # TODO log life type
                    i = i + 1
                    if cols[i] != "life":
                        print("NedL {} '{}'".format(s.key, c))
                        continue
                    i = i + 1

                # find last column
                last_col = num_cols - 2
                last_col_val = cols[last_col]

                if type(last_col_val) is int or last_col_val.isdigit():
                    last_survey = int(last_col_val)
                else:
                    w.owner = last_col_val
                    last_col = last_col - 1
                    last_col_val = cols[last_col]
                    if type(last_col_val) is int or last_col_val.isdigit():
                        last_survey = int(last_col_val)

                last_col = last_col - 1

                regex = re.compile("Special: Strategic Resource")
                if regex.search(c):
                    # SRP world, TODO capture special resource name

                    #desc = ' '.join(cols[3:last_col])
                    last_col = cols.index('Special:') - 2
                    w.srp = float(cols[i+1])
                    last_col = last_col - 1
                else:
                    regex = re.compile("Naturally")
                    if (regex.search(c)):
                        last_col = cols.index('Naturally') - 1
                        w.srp = float(cols[i+1])
                        last_col = last_col - 1
                    else:
                        w.srp = 0

                w.rp = int(cols[i])

                if i != last_col:
                    print("Ned4 {} {} {} {} '{}'".format(s.key, i, last_col, w.rp, c))
            #}

            s.last_survey = last_survey
            s.worlds.append(w)
        #} # foreach line in comments
        # make sure worlds are in pnum order
        s.worlds.sort(key=lambda x:x.pnum)

    def buildStarmap(self, map_sheet):
        self.first_x, last_x = Fots.findX(map_sheet)
        self.first_y, last_y = Fots.findY(map_sheet)

        # walk the map
        for col in range(3, (last_x - self.first_x + 1) * 10):
        #{
            for row in range(4, (last_y - self.first_y + 1) * 10):
                cell = map_sheet.cell(row, col)
                val = cell.value
                if val == None:
                    continue

                s = System()

                s.xoff = self.first_x * 10 + col - 3
                s.yoff = self.first_y * 10 + row - 4

                key = Fots.makeKey(s.xoff, s.yoff)
                s.key = key

                s.grid = cell.fill.start_color.index

                comment = cell.comment
                if type(comment) is not openpyxl.comments.comments.Comment:
                    s.star_type = val
                else:
                    comments = comment.text.split("\n")
                    com_cols = comments[0].split()
                    if key != com_cols[0]:
                        print("Ned error ", key, com_cols[0], file=sys.stderr)
                    s.star_type = com_cols[1]
                    self.parseComments(s, comments)

                self.stars.append(s)
        #}

    def checkSurveys(self, survey_sheet):
        """ check starmap built from comments against survey data """
        # build map from key to system
        starmap = dict()
        for s in self.stars:
            starmap[s.key] = s

        # find first system
        row = 4
        key_col = 2
        key_val = survey_sheet.cell(row, key_col).value
        while key_val == None:
            row = row + 1
            key_val = survey_sheet.cell(row, key_col).value

        pnum_col = 3
        hab_col = 4
        #tag_col = 5
        rp_col = 6
        srp_col = 7
        ssrp_col = 8
        own_col = 16
        while key_val != None:
            pnum_val = survey_sheet.cell(row, pnum_col).value
            if pnum_val is not None and (type(pnum_val) is int or pnum_val.isdigit()):
                # check survey
                sys_key = key_val.split()[0]
                if sys_key in starmap:
                    s = starmap[sys_key]
                    pnum = int(pnum_val) - 1
                    if len(s.worlds) > pnum:
                        tmp_w = World()
                        tmp_w.type = survey_sheet.cell(row, hab_col).value
                        tmp_w.rp = int(survey_sheet.cell(row, rp_col).value)

                        tmp_w.srp = 0
                        srp_val = survey_sheet.cell(row, srp_col).value
                        if srp_val is not None:
                            tmp_w.srp = float(srp_val)

                        ssrp_val = survey_sheet.cell(row, ssrp_col).value
                        if ssrp_val is not None:
                            tmp_w.srp = tmp_w.srp + float(ssrp_val)

                        # TODO resolve owner formula
                        #own_val = survey_sheet.cell(row, own_col).value
                        #if own_val is not None:
                            #tmp_w.owner = own_val

                        compareWorlds(key_val, s.worlds[pnum], tmp_w)
                    else:
                        print("Planet out of range {} {}".format(key_val, pnum))
                        s.worlds.append(tmp_w)
                else:
                    print("Not found {}".format(key_val))

            row = row + 1
            key_val = survey_sheet.cell(row, key_col).value

